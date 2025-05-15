import os
from io import BytesIO
from typing import Union, Dict, List
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Создание логгера
logging.basicConfig(level=logging.INFO)


def safe_float(value):
    """Преобразует значение в тип float, если возможно, или возвращает 0."""
    try:
        return float(value) if value else 0
    except ValueError:
        return 0


def losses_parameters_create_excel_report(data: Union[List[dict], dict]) -> BytesIO:
    """Создаёт Excel отчёт по данным потерь (losses) товаров и возвращает BytesIO объект."""

    # Обработка данных, если пришёл список
    if isinstance(data, list):
        if not data:
            raise ValueError("Пустой список данных")
        data = data[0]  # Используем первый элемент списка

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Потери по товарам"

    # Заголовки
    headers = [
        "Товар",
        "Средняя цена прошлый месяц",
        "Средняя цена прошлая неделя",
        "Средняя цена текущий месяц",
        "Средняя цена текущая неделя",
        "Потери (прошлый месяц к позапрошлому)",
        "Потери (прошлая неделя к позапрошлой)",
        "Потери (текущий месяц к прошлому)"
    ]

    # Добавление заголовков в Excel
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin"))

    # Заполнение данных
    total_loss_1 = total_loss_2 = total_loss_3 = 0
    for row_num, item in enumerate(data.get("data", []), start=2):
        avg_price_last_month = safe_float(item.get('avg_price_last_month'))
        avg_price_last_week = safe_float(item.get('avg_price_last_week'))
        avg_price_current_month = safe_float(item.get('avg_price_current_month'))
        avg_price_current_week = safe_float(item.get('avg_price_current_week'))

        loss_1 = safe_float(item.get('losses_last_month_to_month_before_last'))
        loss_2 = safe_float(item.get('losses_last_week_to_week_before_last'))
        loss_3 = safe_float(item.get('losses_current_month_to_last'))

        total_loss_1 += loss_1
        total_loss_2 += loss_2
        total_loss_3 += loss_3

        row = [
            item.get("label", ""),
            avg_price_last_month,
            avg_price_last_week,
            avg_price_current_month if item.get("avg_price_current_month") else "-",
            avg_price_current_week if item.get("avg_price_current_week") else "-",
            loss_1,
            loss_2,
            loss_3
        ]

        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)

        # Цвета для ячеек
        if loss_3 > 1000:
            ws.cell(row=row_num, column=8).font = Font(color="FF0000")
        if avg_price_current_month > 500:
            ws.cell(row=row_num, column=4).font = Font(color="008000")

    # Добавление итогов
    total_row = len(data.get("data", [])) + 2
    ws.cell(row=total_row, column=5, value="ИТОГО:")
    ws.cell(row=total_row, column=6, value=total_loss_1)
    ws.cell(row=total_row, column=7, value=total_loss_2)
    ws.cell(row=total_row, column=8, value=total_loss_3)

    # Применение стилей
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin"))

    # Сохранение в буфер и возврат объекта BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def losses_only_negative_create_excel_report(data: Union[List[dict], dict]) -> BytesIO:
    """Создаёт Excel отчёт по данным потерь (losses) товаров и возвращает BytesIO объект."""

    # Обработка данных, если пришёл список
    if isinstance(data, list):
        if not data:
            raise ValueError("Пустой список данных")
        data = data[0]  # Используем первый элемент списка

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Потери по товарам"

    # Заголовки
    headers = [
        "Товар",
        "Средняя цена прошлый месяц",
        "Средняя цена прошлая неделя",
        "Средняя цена текущий месяц",
        "Средняя цена текущая неделя",
        "Потери (прошлый месяц к позапрошлому)",
        "Потери (прошлая неделя к позапрошлой)",
        "Потери (текущий месяц к прошлому)"
    ]

    # Добавление заголовков в Excel
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin"))

    # Заполнение данных
    total_loss_1 = total_loss_2 = total_loss_3 = 0
    for row_num, item in enumerate(data.get("data", []), start=2):
        avg_price_last_month = safe_float(item.get('avg_price_last_month'))
        avg_price_last_week = safe_float(item.get('avg_price_last_week'))
        avg_price_current_month = safe_float(item.get('avg_price_current_month'))
        avg_price_current_week = safe_float(item.get('avg_price_current_week'))

        loss_1 = safe_float(item.get('losses_last_month_to_month_before_last'))
        loss_2 = safe_float(item.get('losses_last_week_to_week_before_last'))
        loss_3 = safe_float(item.get('losses_current_month_to_last'))

        total_loss_1 += loss_1
        total_loss_2 += loss_2
        total_loss_3 += loss_3

        # Проверяем только те потери, которые больше нуля
        if loss_1 > 0 or loss_2 > 0 or loss_3 > 0:
            row = [
                item.get("label", ""),
                avg_price_last_month,
                avg_price_last_week,
                avg_price_current_month if avg_price_current_month else "-",
                avg_price_current_week if avg_price_current_week else "-",
                loss_1 if loss_1 > 0 else "-",
                loss_2 if loss_2 > 0 else "-",
                loss_3 if loss_3 > 0 else "-"
            ]

            # Заполняем ячейки
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

            # Цвета для ячеек
            if loss_3 > 1000:
                ws.cell(row=row_num, column=8).font = Font(color="FF0000")
            if avg_price_current_month > 500:
                ws.cell(row=row_num, column=4).font = Font(color="008000")

    # Добавление итогов
    total_row = len(data.get("data", [])) + 2
    ws.cell(row=total_row, column=5, value="ИТОГО:")
    ws.cell(row=total_row, column=6, value=total_loss_1)
    ws.cell(row=total_row, column=7, value=total_loss_2)
    ws.cell(row=total_row, column=8, value=total_loss_3)

    # Применение стилей
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin"))

    # Сохранение в буфер и возврат объекта BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer
