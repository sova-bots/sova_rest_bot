import json
import logging

import pandas as pd
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from aiogram import Router, types, F
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import os

from aiogram.filters import Command

# Helper functions (no change from previous code)
def get_first_non_null(*values):
    for value in values:
        if value is not None:
            return value
    return "-"

foodcost_of_products_storehouse_excel_router = Router()

# Helper functions (your previously provided functions)
def get_first_non_null(*values):
    for value in values:
        if value is not None:
            return value
    return "-"

def calculate_percentage_change(old_value, new_value):
    if isinstance(old_value, (int, float)) and isinstance(new_value, (int, float)) and old_value != 0:
        return round(((new_value - old_value) / old_value) * 100, 2)
    return None

def calculate_monthly_differences(store):
    diff_price = store.get("food_cost_dynamics_day")
    diff_price2 = store.get("food_cost_dynamics_week")
    diff_price3 = store.get("food_cost_dynamics_month")

    diff_price_1 = get_first_non_null(diff_price, diff_price2, diff_price3)

    if isinstance(diff_price2, (int, float)) and isinstance(diff_price, (int, float)):
        diff_price_1_day = calculate_percentage_change(diff_price, diff_price2)
    else:
        diff_price_1_day = "-"

    if isinstance(diff_price3, (int, float)) and isinstance(diff_price2, (int, float)):
        diff_price_2_week = calculate_percentage_change(diff_price2, diff_price3)
    else:
        diff_price_2_week = "-"

    return diff_price_1, diff_price_1_day, diff_price_2_week

def safe_format(value):
    if value is None:
        return "-"
    if isinstance(value, (int, float)):
        return f"{value:,.2f}%"  # Add percentage formatting
    return str(value)

def create_excel_report(data, output_file="food_cost_server_report.xlsx"):
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Food Cost Report"

    # Headers for the table
    headers = ["Название", "Фудкост", "Динамика день", "Динамика неделя", "Динамика месяц"]

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Prepare table data
    table_data = []
    for product in data["data"]:
        food_cost_1, food_cost_1_day, food_cost_2_week = calculate_monthly_differences(product)
        row = [
            product["label"],
            safe_format(product['food_cost']),
            safe_format(food_cost_1_day),
            safe_format(food_cost_1),
            safe_format(food_cost_2_week),
        ]
        table_data.append(row)

    # Write data to the Excel sheet
    for row_num, row_data in enumerate(table_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    column_widths = [max(len(str(cell.value)) for cell in col) for col in zip(*ws.iter_cols())]
    for i, col_width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = col_width + 2

    # Save the workbook to the specified file
    wb.save(output_file)
    print(f"Excel-файл успешно сохранен: {output_file}")

def load_json_data(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return data
    except Exception as e:
        print(f"Ошибка загрузки данных JSON из {file_path}: {e}")
        return None


@foodcost_of_products_storehouse_excel_router.callback_query(F.data == "format_excel_food_cost_dynamics")
async def generate_excel_report_callback(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer("Формирую Excel отчёт по себестоимости продуктов на складе...")

    # Извлекаем тип отчёта из callback_data
    report_type = callback_query.data.split("_")[-1]

    # Загрузка данных из JSON-файла
    filepath = r"C:\WORK\sova_rest_bot\sova_rest_bot-master\files\jsons_for_reports\food-cost-dish_server_data_example.json"
    foodcost_data = load_json_data(filepath)  # Используем функцию для загрузки данных

    if not foodcost_data:
        await callback_query.message.answer("Ошибка при загрузке данных для отчёта.")
        return

    # Генерация Excel
    try:
        filename = f"foodcost_storehouse_{report_type}.xlsx"
        create_excel_report(foodcost_data, filename)  # Используем функцию для создания Excel
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        await callback_query.message.answer(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        with open(filename, "rb") as file:
            file_data = file.read()

        input_file = BufferedInputFile(file_data, filename=f"foodcost_storehouse_{report_type}.xlsx")
        await callback_query.message.answer_document(
            input_file,
            caption=f"Ваш отчёт по себестоимости продуктов на складе (тип: {report_type}):"
        )
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")
        await callback_query.message.answer("Ошибка при отправке отчёта.")
    finally:
        # Удаляем временный файл после отправки
        if os.path.exists(filename):
            os.remove(filename)