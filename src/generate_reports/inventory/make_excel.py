import logging
import os

import openpyxl
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from openpyxl.styles import PatternFill, Alignment, Font
from io import BytesIO
import json
from aiogram import types, F
from aiogram import Router
from aiogram.types import FSInputFile, BufferedInputFile


# Define safe_float function
def safe_float(value):
    try:
        return float(value) if value else 0
    except (ValueError, TypeError):
        return 0


# Define the function that generates the Excel report
def create_excel_report(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Инвентаризация товаров"

    header = ["Магазин", "Недостача", "Недостача (%)", "Излишки", "Излишки (%)", "Себестоимость"]
    ws.append(header)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center")
    font_bold = Font(bold=True)

    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.alignment = center_align
        cell.font = font_bold
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for i, store in enumerate(data["data"], start=2):
        store_name = store["label"]
        shortage = safe_float(store.get('shortage', 0))
        shortage_percent = safe_float(store.get('shortage_percent', 0))
        surplus = safe_float(store.get('surplus', 0))
        surplus_percent = safe_float(store.get('surplus_percent', 0))
        cost_price = safe_float(store.get('cost_price', 0))

        ws.append([
            store_name,
            f"{shortage:,.2f}",
            f"{shortage_percent:,.2f}%",
            f"{surplus:,.2f}",
            f"{surplus_percent:,.2f}%",
            f"{cost_price:,.2f}"
        ])

        if shortage_percent > 2:
            ws.cell(row=i, column=2).fill = red_fill
        if surplus_percent > 3:
            ws.cell(row=i, column=4).fill = green_fill

        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            cell.alignment = center_align

    ws.column_dimensions['A'].width = 30
    for col in range(2, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


# Create the router for handling inventory Excel report
inventory_excel_router = Router()


# Загрузка данных
def load_json_data(filepath: str) -> dict:
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        logging.error(f"Ошибка при чтении файла {filepath}: {e}")
        return {}


@inventory_excel_router.callback_query(F.data == 'inventory_excel')
async def generate_inventory_report_callback(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer("Формирую Excel отчёт по инвентаризации...")

    # Загрузка данных из JSON-файла
    filepath = r"C:\WORK\sova_rest_bot\sova_rest_bot-master\files\jsons_for_reports\inventory_store_example.json"
    inventory_data = load_json_data(filepath)  # Используем функцию для загрузки данных

    if not inventory_data:
        await callback_query.message.answer("Ошибка при загрузке данных для отчёта.")
        return

    # Генерация Excel
    try:
        excel_buffer = create_excel_report(inventory_data)  # Получаем BytesIO объект
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        await callback_query.message.answer(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        # Используем данные из BytesIO для создания BufferedInputFile
        input_file = BufferedInputFile(excel_buffer.getvalue(), filename="inventory_store_report.xlsx")
        await callback_query.message.answer_document(
            input_file,
            caption="Ваш отчёт по инвентаризации готов!"
        )
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")
        await callback_query.message.answer("Ошибка при отправке отчёта.")
    finally:
        # Закрываем BytesIO объект
        excel_buffer.close()