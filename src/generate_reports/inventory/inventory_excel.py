import logging
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from io import BytesIO
import json
from aiogram import Router

# Define safe_float function
def safe_float(value):
    try:
        return float(value) if value else 0
    except (ValueError, TypeError):
        return 0


def create_excel_report(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Инвентаризация по складам"

    # Удалён столбец "Себестоимость"
    header = ["Магазин", "Недостача", "Недостача (% от с/c)", "Излишки", "Излишки (% от с/с)"]
    ws.append(header)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center")
    font_bold = Font(bold=True)

    for col in range(1, 6):  # Всего 5 столбцов
        cell = ws.cell(row=1, column=col)
        cell.alignment = center_align
        cell.font = font_bold
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    store_data = data[0]["data"]

    for i, store in enumerate(store_data, start=2):
        store_name = store["label"]
        shortage = safe_float(store.get('shortage', 0))
        shortage_percent = safe_float(store.get('shortage_percent', 0))
        surplus = safe_float(store.get('surplus', 0))
        surplus_percent = safe_float(store.get('surplus_percent', 0))

        ws.append([
            store_name,
            f"{shortage:,.2f}",
            f"{shortage_percent:,.2f}%",
            f"{surplus:,.2f}",
            f"{surplus_percent:,.2f}%"
        ])

        if shortage_percent > 2:
            ws.cell(row=i, column=2).fill = red_fill
        if surplus_percent > 3:
            ws.cell(row=i, column=4).fill = green_fill

        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            cell.alignment = center_align

    ws.column_dimensions['A'].width = 30
    for col in range(2, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


