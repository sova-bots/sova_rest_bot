import json
import logging
import os
from aiogram.types import CallbackQuery, BufferedInputFile, FSInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from aiogram import Router, F
from src.basic.revenue_analysis.graphics_for_pdf import load_revenue_data

# Initialize the routers
analys_revenue_excel_router = Router()

# Function to create the revenue excel file
def create_revenue_excel(data: dict, filename: str):
    """Creates an Excel file with revenue analysis"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ выручки"

    # Create style for formatting numbers with thousands separator
    number_style = NamedStyle(name="number_style", number_format="#.##0")
    wb.add_named_style(number_style)

    # Column headers
    headers = [
        "Заведение", "Выручка", "Выручка за неделю", "Выручка за месяц",
        "Выручка за год", "Динамика выручки за неделю (%)",
        "Динамика выручки за месяц (%)", "Динамика выручки за год (%)",
        "Прогноз выручки"
    ]
    ws.append(headers)

    # Bold font for headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Fill data
    for item in data["data"]:
        row = [
            item["label"],
            item["revenue"],
            item["revenue_week"],
            item["revenue_month"],
            item["revenue_year"],
            item["revenue_dynamics_week"],
            item["revenue_dynamics_month"],
            item["revenue_dynamics_year"],
            item["revenue_forecast"]
        ]
        ws.append(row)

    # Add total row
    total_row = [
        data["sum"]["label"],
        data["sum"]["revenue"],
        data["sum"]["revenue_week"],
        data["sum"]["revenue_month"],
        data["sum"]["revenue_year"],
        data["sum"]["revenue_dynamics_week"],
        data["sum"]["revenue_dynamics_month"],
        data["sum"]["revenue_dynamics_year"],
        data["sum"]["revenue_forecast"]
    ]
    ws.append(total_row)

    # Apply number formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=9):
        for cell in row:
            cell.style = number_style

    # Save file
    wb.save(filename)
    return filename


@analys_revenue_excel_router.callback_query(F.data == "revenue_analysis_excel")
async def handle_format_excel(callback_query: CallbackQuery):
    """Обработчик для кнопки 'Сформировать Excel отчёт'."""
    # Отвечаем на callback_query, чтобы убрать "часики" у кнопки
    await callback_query.answer("Формирую Excel отчёт...")

    # Извлекаем тип отчёта из callback_data
    report_type = callback_query.data.split("_")[-1]

    # Загрузка данных из JSON-файла
    filepath = r"C:\WORK\sova_rest_bot\sova_rest_bot-master\files\jsons_for_reports\revenue analys.json"
    revenue_data = load_revenue_data(filepath)

    if not revenue_data:
        await callback_query.message.answer("Ошибка при загрузке данных для отчёта.")
        return

    # Генерация Excel
    try:
        filename = create_revenue_excel(revenue_data, filename=f"revenue_analysis_{report_type}.xlsx")
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        await callback_query.message.answer(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        with open(filename, "rb") as file:
            file_data = file.read()

        input_file = BufferedInputFile(file_data, filename=f"revenue_analysis_{report_type}.xlsx")
        await callback_query.message.answer_document(input_file, caption=f"Ваш отчёт по анализу выручки (тип: {report_type}):")
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")
        await callback_query.message.answer("Ошибка при отправке отчёта.")


