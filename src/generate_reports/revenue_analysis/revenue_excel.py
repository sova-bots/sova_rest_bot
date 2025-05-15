import logging
import os
import json
from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
import tempfile
from aiogram import F, Router, types
from aiogram.types import CallbackQuery, FSInputFile
from aiogram.fsm.context import FSMContext
from src.analytics.handlers.msg.headers import make_header
from src.analytics.handlers.types.msg_data import MsgData
from typing import Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Union
import logging


# Цвета для заливки ячеек
POSITIVE_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Зеленый
NEGATIVE_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Красный
HEADER_FILL = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")  # Серый для заголовков

# Цвета для текста
POSITIVE_FONT = Font(color="006400")  # Зеленый
NEGATIVE_FONT = Font(color="B71C1C")  # Красный
DEFAULT_FONT = Font(color="000000")  # Черный


def format_currency(value):
    try:
        if value is None:
            return "—"
        if isinstance(value, str):
            value = value.lstrip("0") or "0"
            value = int(value)
        elif not isinstance(value, int):
            value = int(value)
        return f"{value:,}".replace(",", " ")
    except Exception as e:
        logging.error(f"Ошибка форматирования валюты: {e}")
        return str(value)


def format_percentage(value: Union[int, float, None]) -> str:
    try:
        if value is None:
            return "—"
        return f"{value}%"
    except Exception as e:
        logging.error(f"Ошибка форматирования процентов: {e}")
        return str(value) + "%"

def load_revenue_data(filepath):
    """Загружает данные из JSON файла"""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"Ошибка при загрузке данных из файла {filepath}: {e}")
        return None


def create_revenue_excel(data: list, filename: str):
    """Создаёт Excel файл с анализом выручки"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ выручки"

    # Создание стиля для форматирования чисел с разделителем тысяч
    number_style = NamedStyle(name="number_style", number_format="#,##0.00")
    wb.add_named_style(number_style)

    # Заголовки столбцов (обновлённые)
    headers = [
        "Заведение", "Выручка",
        "Динамика неделя",
        "Динамика месяц",
        "Динамика год",
        "Прогноз"
    ]
    ws.append(headers)

    # Полужирный шрифт и заливка для заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    # Получаем первый элемент списка данных
    data = data[0]  # Используем только первый элемент списка

    # Заполнение данными
    for item in data["data"]:
        row = [
            item["label"],
            f"{item['revenue']:.2f} ",
            f"{item['revenue_dynamics_week']:.2f} %",
            f"{item['revenue_dynamics_month']:.2f} %",
            f"{item['revenue_dynamics_year']:.2f} %",
            f"{item['revenue_forecast']:.2f} "
        ]
        ws.append(row)

        # Применение заливки и цвета текста для динамики
        for col_idx, key in zip([3, 4, 5], [
            "revenue_dynamics_week",
            "revenue_dynamics_month",
            "revenue_dynamics_year"
        ]):
            dynamics_value = item[key]
            cell = ws.cell(row=ws.max_row, column=col_idx + 1)
            if dynamics_value > 0:
                cell.fill = POSITIVE_FILL
                cell.font = POSITIVE_FONT
            else:
                cell.fill = NEGATIVE_FILL
                cell.font = NEGATIVE_FONT
            cell.value = f"{dynamics_value:.2f} %"

    # Строка итогов
    total_row = [
        data["sum"]["label"],
        f"{data['sum']['revenue']:.2f} ",
        f"{data['sum']['revenue_dynamics_week']:.2f} %",
        f"{data['sum']['revenue_dynamics_month']:.2f} %",
        f"{data['sum']['revenue_dynamics_year']:.2f} %",
        f"{data['sum']['revenue_forecast']:.2f} "
    ]
    ws.append(total_row)

    # Форматирование чисел
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=6):
        for cell in row:
            cell.style = number_style

    # Стиль для итоговой строки
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)

    # Сохранение
    wb.save(filename)
    return filename


def send_excel_report(filepath):
    """Основная функция для создания и отправки Excel отчёта"""
    revenue_data = load_revenue_data(filepath)
    if not revenue_data:
        logging.error("Ошибка при загрузке данных для отчёта.")
        return

    # Получаем первый элемент списка, который содержит словарь
    revenue_data = revenue_data[0]

    # Генерация Excel файла
    try:
        filename = create_revenue_excel(revenue_data, filename=f"revenue_analysis.xlsx")
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        with open(filename, "rb") as file:
            file_data = file.read()

        input_file = BufferedInputFile(file_data, filename=f"revenue_analysis.xlsx")

        logging.info(f"Отчёт успешно отправлен: {filename}")
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")



revenue_excel_report_router = Router()

# Обработчик для кнопки Excel отчета
@revenue_excel_report_router.callback_query(F.data == "report:send_excel_report")
async def handle_send_excel_report(callback: CallbackQuery, state: FSMContext):
    msg_data = MsgData(
        msg=callback.message,
        state=state,
        tgid=callback.from_user.id
    )
    await send_excel_report(callback, msg_data)

async def send_excel_report(callback: CallbackQuery, msg_data: MsgData) -> None:
    state_data = await msg_data.state.get_data()
    report_format = "excel"

    header = await make_header(msg_data)
    text = f"{header}\n\n📎 Excel отчёт прикреплён."

    try:
        json_file_path = state_data.get("report:json_file_path")

        # Автогенерация JSON, если путь не найден
        if not json_file_path or not os.path.exists(json_file_path):
            json_data = state_data.get("report:json_data")
            if not json_data:
                await callback.message.answer("Ошибка: нет данных для генерации Excel.")
                return

            type_data = state_data.get("report:type", "revenue")
            period = state_data.get("report:period", "weekly")
            file_name = f"{type_data}_json_{period}.json"

            json_file_path = os.path.join(tempfile.gettempdir(), file_name)
            with open(json_file_path, "w", encoding="utf-8") as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)

            # Обновим state
            state_data["report:json_file_path"] = json_file_path
            await msg_data.state.update_data(state_data)

        # Генерация Excel
        excel_filename = os.path.join(tempfile.gettempdir(), "revenue_report.xlsx")
        revenue_data = load_revenue_data(json_file_path)
        create_revenue_excel(revenue_data, excel_filename)

        await callback.message.answer_document(
            document=FSInputFile(excel_filename),
            caption=text,
        )

    except Exception as e:
        await callback.message.answer(f"Ошибка при создании отчёта: {e}")
    finally:
        # Удаление временных файлов
        if "json_file_path" in locals() and os.path.exists(json_file_path):
            os.remove(json_file_path)
        if "excel_filename" in locals() and os.path.exists(excel_filename):
            os.remove(excel_filename)


logging.basicConfig(level=logging.INFO)

def safe_float(value):
    try:
        return float(value) if value not in [None, ""] else 0
    except:
        return 0

def format_rub(val):
    try:
        return f"{int(val):,}".replace(",", " ") + " "
    except:
        return str(val) or "—"

def format_percent(val):
    try:
        return f"{float(val):.2f} %"
    except:
        return "—"

def revenue_parameters_create_excel_report_analysis(data: Union[List[dict], dict]) -> BytesIO:
    if isinstance(data, list):
        if not data:
            raise ValueError("Пустой список данных")
        data = data[:9]  # максимум 9 блоков
    else:
        raise ValueError("Ожидался список JSON-блоков")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # удаляем стандартный пустой лист

    # Общие стили
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_font = Font(color="FF0000")
    green_font = Font(color="008000")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin"))

    # Конфигурация листов
    configs = [
        {
            "title": "Гости-чеки",
            "columns": [
                ("Подразделение", "label"),
                ("Гости", "guests"),
                ("Динамика неделя", "guests_dynamics_week"),
                ("Динамика месяц", "guests_dynamics_month"),
                ("Динамика год", "guests_dynamics_year"),
                ("Чеки", "checks"),
                ("Динамика неделя", "checks_dynamics_week"),
                ("Динамика месяц", "checks_dynamics_month"),
                ("Динамика год", "checks_dynamics_year"),
            ],
            "format": ["text", "int", "percent", "percent", "percent", "int", "percent", "percent", "percent"]
        },
        {
            "title": "Средний чек",
            "columns": [
                ("Подразделение", "label"),
                ("Средний чек", "avg_check"),
                ("Динамика неделя", "avg_check_dynamics_week"),
                ("Динамика месяц", "avg_check_dynamics_month"),
                ("Динамика год", "avg_check_dynamics_year"),
            ],
            "format": ["text", "rub", "percent", "percent", "percent"]
        },
        {
            "title": "Выручка",
            "columns": [
                ("Подразделение", "label"),
                ("Выручка", "revenue"),
                ("Динамика неделя", "revenue_dynamics_week"),
                ("Динамика месяц", "revenue_dynamics_month"),
                ("Динамика год", "revenue_dynamics_year"),
                ("Прогноз", "revenue_forecast"),
            ],
            "format": ["text", "rub", "percent", "percent", "percent", "rub"]
        },
        {
            "title": "Выручка по направлениям",
            "columns": [
                ("Направление", "label"),
                ("Выручка", "revenue"),
                ("Динамика неделя", "revenue_dynamics_week"),
                ("Динамика месяц", "revenue_dynamics_month"),
                ("Динамика год", "revenue_dynamics_year"),
            ],
            "format": ["text", "rub", "percent", "percent", "percent"]
        },
        {
            "title": "Выручка по блюдам",
            "columns": [
                ("Блюдо", "label"),
                ("Выручка", "revenue"),
                ("Динамика неделя", "revenue_dynamics_week"),
                ("Динамика месяц", "revenue_dynamics_month"),
                ("Динамика год", "revenue_dynamics_year"),
            ],
            "format": ["text", "rub", "percent", "percent", "percent"]
        },
        {
            "title": "Выручка по времени посещения",
            "columns": [
                ("Время", "label"),
                ("Выручка", "revenue"),
                ("Динамика неделя", "revenue_dynamics_week"),
                ("Динамика месяц", "revenue_dynamics_month"),
                ("Динамика год", "revenue_dynamics_year"),
            ],
            "format": ["text", "rub", "percent", "percent", "percent"]
        },
        {
            "title": "Выручка по ценовым сегментам",
            "columns": [
                ("Ценовой сегмент", "label"),
                ("Выручка", "revenue"),
                ("Динамика неделя", "revenue_dynamics_week"),
                ("Динамика месяц", "revenue_dynamics_month"),
                ("Динамика год", "revenue_dynamics_year"),
            ],
            "format": ["text", "rub", "percent", "percent", "percent"]
        },
        {
            "title": "Выручка по дням недели",
            "columns": [
                ("День", "label"),
                ("Выручка", "revenue"),
                ("Динамика неделя", "revenue_dynamics_week"),
                ("Динамика месяц", "revenue_dynamics_month"),
                ("Динамика год", "revenue_dynamics_year"),
            ],
            "format": ["text", "rub", "percent", "percent", "percent"]
        },
        {
            "title": "Анализ работы официантов",
            "columns": [
                ("Сотрудник", "label"),
                ("Выручка", "revenue"),
                ("Средняя выручка", "avg_revenue"),
                ("Чек", "avg_checks"),
                ("Глубина чека", "depth"),
                ("Потенциал", "potential"),
            ],
            "format": ["text", "rub", "rub", "int", "float", "rub"]
        },
    ]

    for i, block in enumerate(data):
        cfg = configs[i]
        ws = wb.create_sheet(title=cfg["title"])
        columns = cfg["columns"]
        formats = cfg["format"]

        # Заголовки
        for col_num, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Данные
        for row_num, item in enumerate(block.get("data", []), start=2):
            for col_num, ((_, key), fmt) in enumerate(zip(columns, formats), start=1):
                val = item.get(key)
                cell = ws.cell(row=row_num, column=col_num)
                if fmt == "rub":
                    cell.value = format_rub(val)
                elif fmt == "percent":
                    cell.value = format_percent(val)
                    if safe_float(val) > 0:
                        cell.font = green_font
                    elif safe_float(val) < 0:
                        cell.font = red_font
                elif fmt == "int":
                    try:
                        cell.value = int(val)
                    except:
                        cell.value = val or 0
                elif fmt == "float":
                    try:
                        cell.value = f"{float(val):.2f}"
                    except:
                        cell.value = "—"
                else:
                    cell.value = val or "—"

                cell.alignment = center_align
                cell.border = thin_border

        # Итоги
        total = block.get("sum", {})
        total_row = ws.max_row + 1
        for col_num, ((_, key), fmt) in enumerate(zip(columns, formats), start=1):
            cell = ws.cell(row=total_row, column=col_num)
            val = total.get(key)
            if fmt == "rub":
                cell.value = format_rub(val)
            elif fmt == "percent":
                cell.value = format_percent(val)
            elif fmt == "int":
                try:
                    cell.value = int(val)
                except:
                    cell.value = 0
            elif fmt == "float":
                try:
                    cell.value = f"{float(val):.2f}"
                except:
                    cell.value = "—"
            else:
                cell.value = val or "Итого" if col_num == 1 else "—"
            cell.font = bold_font
            cell.fill = gray_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Автоширина
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_len + 2

    # Сохраняем в BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
