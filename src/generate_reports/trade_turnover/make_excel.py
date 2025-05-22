import json
import logging
import os
from aiogram.types import CallbackQuery, BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from aiogram import Router, F

trade_turnover_excel_report_router = Router()

def create_excel_report(data: dict, filename: str):
    """Создаёт Excel-файл с анализом"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ данных"

    number_style = NamedStyle(name="number_style", number_format="#,##0")
    wb.add_named_style(number_style)

    headers = [
        "Заведение", "Расход в день", "Оборачиваемость (дни)",
        "Динамика оборачиваемости за неделю", "Динамика за месяц", "Динамика за год",
        "Оборачиваемость за неделю", "Оборачиваемость за месяц", "Оборачиваемость за год",
        "Остаток на конец периода"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in data["data"]:
        row = [
            item["label"], item["expense_day"], item["turnover_in_days"],
            item["turnover_in_days_dynamic_week"], item["turnover_in_days_dynamic_month"],
            item["turnover_in_days_dynamic_year"], item["turnover_in_days_week"],
            item["turnover_in_days_month"], item["turnover_in_days_year"],
            item["remainder_end"]
        ]
        ws.append(row)

    total_row = [
        data["sum"]["label"], data["sum"]["expense_day"], data["sum"]["turnover_in_days"],
        data["sum"]["turnover_in_days_dynamic_week"], data["sum"]["turnover_in_days_dynamic_month"],
        data["sum"]["turnover_in_days_dynamic_year"], data["sum"]["turnover_in_days_week"],
        data["sum"]["turnover_in_days_month"], data["sum"]["turnover_in_days_year"],
        data["sum"]["remainder_end"]
    ]
    ws.append(total_row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=10):
        for cell in row:
            cell.style = number_style

    wb.save(filename)
    return filename

def load_revenue_data(filepath: str) -> dict:
    """Загружает данные из JSON-файла."""
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        logging.error(f"Ошибка при чтении файла {filepath}: {e}")
        return {}

@trade_turnover_excel_report_router.callback_query(F.data == "format_excel_turnover")
async def handle_excel_request(callback_query: CallbackQuery):
    """Обработчик для кнопки 'Сформировать Excel отчёт по товарообороту'."""
    # Отвечаем на callback_query, чтобы убрать "часики" у кнопки
    await callback_query.answer("Формирую Excel отчёт по товарообороту...")

    # Извлекаем тип отчёта из callback_data
    report_type = callback_query.data.split("_")[-1]

    # Загрузка данных из JSON-файла
    filepath = r"C:\WORK\sova_rest_bot\sova_rest_bot-master\files\jsons_for_reports\turnover-store_example.json"
    turnover_data = load_revenue_data(filepath)  # Используем ту же функцию для загрузки данных

    if not turnover_data:
        await callback_query.message.answer("Ошибка при загрузке данных для отчёта.")
        return

    # Генерация Excel
    try:
        filename = create_excel_report(turnover_data, filename=f"turnover_{report_type}.xlsx")
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        await callback_query.message.answer(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        with open(filename, "rb") as file:
            file_data = file.read()

        input_file = BufferedInputFile(file_data, filename=f"turnover_{report_type}.xlsx")
        await callback_query.message.answer_document(input_file, caption=f"Ваш отчёт по товарообороту (тип: {report_type}):")
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")
        await callback_query.message.answer("Ошибка при отправке отчёта.")
    finally:
        # Удаляем временный файл после отправки
        if os.path.exists(filename):
            os.remove(filename)