import json
import logging
import os

from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery, BufferedInputFile, FSInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from aiogram import Router, F, types
import json
import logging
import os
from aiogram import Router, F
from aiogram.types import CallbackQuery, BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Initialize router
forecasting_losses_excel_router = Router()

# Path to the JSON file containing the data
json_file_path = r'C:\\WORK\\sova_rest_bot\\sova_rest_bot-master\\files\\jsons_for_reports\\loss-forecast_data_example.json'


def check_font_path(font_path):
    if not os.path.exists(font_path):
        raise FileNotFoundError(f"Шрифт не найден по пути: {font_path}")


def get_first_non_null(*values):
    for value in values:
        if value is not None:
            return value
    return "-"


def calculate_monthly_differences(store):
    """
    Вычисляет разницы между месяцами для магазина.
    Возвращает кортеж: (diff_price_1, diff_price_1_month, diff_price_2_month)
    """
    # Получаем значения разниц цен
    diff_price = store.get("diff_price")
    diff_price2 = store.get("diff_price2")
    diff_price3 = store.get("diff_price3")
    diff_price4 = store.get("diff_price4")

    # Первое изменение цены (первое не-null значение)
    diff_price_1 = get_first_non_null(diff_price, diff_price2, diff_price3, diff_price4)

    # Разница между 1 и 2 месяцами (diff_price3 - diff_price2)
    if isinstance(diff_price2, (int, float)) and isinstance(diff_price3, (int, float)):
        diff_price_1_month = round(diff_price3 - diff_price2, 2)
    else:
        diff_price_1_month = "-"

    # Разница между 2 и 3 месяцами (diff_price4 - diff_price3)
    if isinstance(diff_price3, (int, float)) and isinstance(diff_price4, (int, float)):
        diff_price_2_month = round(diff_price4 - diff_price3, 2)
    else:
        diff_price_2_month = "-"

    return diff_price_1, diff_price_1_month, diff_price_2_month


def load_json_data(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return data
    except Exception as e:
        print(f"Ошибка загрузки данных JSON из {file_path}: {e}")
        return None


def create_excel_with_table(data, output_file="loss_forecast_report.xlsx"):
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Прогнозирование потерь"

    # Заголовок таблицы
    headers = ["Магазин", "Прогноз", "Первое изменение цены", "Изменение (1 месяц)", "Изменение (2 месяца)"]
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Применяем стили к заголовкам
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Сортировка магазинов по убыванию прогноза
    sorted_data = sorted(data["data"], key=lambda store: store.get("forecast", 0), reverse=True)

    for store in sorted_data:
        forecast = store.get("forecast", "-")

        # Вычисляем разницы между месяцами
        diff_price_1, diff_price_1_month, diff_price_2_month = calculate_monthly_differences(store)

        # Пропускаем магазин, если нет данных о разнице цен
        if diff_price_1 == "-" and diff_price_1_month == "-" and diff_price_2_month == "-":
            continue

        # Добавляем строку в таблицу
        row = [
            store["label"],
            f"{forecast:,.2f}" if isinstance(forecast, (int, float)) else forecast,
            f"{diff_price_1:,.2f}" if isinstance(diff_price_1, (int, float)) else "-",
            f"{diff_price_1_month:,.2f}" if isinstance(diff_price_1_month, (int, float)) else "-",
            f"{diff_price_2_month:,.2f}" if isinstance(diff_price_2_month, (int, float)) else "-"
        ]
        ws.append(row)

    # Применяем стили к данным
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Автоматическое выравнивание ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(output_file)
    print(f"Excel-файл успешно сохранен: {output_file}")


# Загрузка данных из JSON
def load_json_data(filepath: str) -> dict:
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        logging.error(f"Ошибка при чтении файла {filepath}: {e}")
        return {}


@forecasting_losses_excel_router.callback_query(F.data == "format_excel_loss_forecast")
async def handle_forecasting_losses_excel(callback_query: types.CallbackQuery, state: FSMContext):
    """Обработчик для кнопки 'Сформировать Excel отчёт по прогнозированию потерь'."""
    await callback_query.answer("Формирую Excel отчёт по прогнозированию потерь...")

    # Загрузка данных из JSON-файла
    filepath = r"C:\WORK\sova_rest_bot\sova_rest_bot-master\files\jsons_for_reports\loss_forecast.json"
    loss_data = load_json_data(filepath)  # Используем функцию для загрузки данных

    if not loss_data:
        await callback_query.message.answer("Ошибка при загрузке данных для отчёта.")
        return

    # Генерация Excel
    try:
        filename = "loss_forecast_report.xlsx"
        create_excel_with_table(loss_data, filename)  # Используем функцию для создания Excel
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        await callback_query.message.answer(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        with open(filename, "rb") as file:
            file_data = file.read()

        input_file = BufferedInputFile(file_data, filename="loss_forecast_report.xlsx")
        await callback_query.message.answer_document(
            input_file,
            caption="Ваш отчёт по прогнозированию потерь:"
        )
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")
        await callback_query.message.answer("Ошибка при отправке отчёта.")
    finally:
        # Удаляем временный файл после отправки
        if os.path.exists(filename):
            os.remove(filename)