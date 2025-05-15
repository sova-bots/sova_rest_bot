from typing import Dict, Union, List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


# Цвета
PASTEL_RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
PASTEL_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")


def create_excel_report_for_turnover(data: Union[Dict, List[Dict]]) -> BytesIO:
    if isinstance(data, list):
        data = data[0] if data else {}

    excel_buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Оборачиваемость остатков"

    headers = [
        "Точка",
        "Остаток на \nконец \nпериода",
        "Остаток \n на конец \nпериода в днях",
        "Динамика \nнеделя",
        "Динамика \nмесяц",
        "Динамика \nгод"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    def create_row(entry: Dict, label: str = None) -> list:
        return [
            f"{label}" if label else entry.get("label", ""),
            f"{safe_float(entry.get('remainder_end')):,.0f}".replace(',', ' '),
            f"{safe_float(entry.get('turnover_in_days')):.1f}",
            f"{safe_float(entry.get('turnover_in_days_dynamic_week')):.1f}",
            f"{safe_float(entry.get('turnover_in_days_dynamic_month')):.1f}",
            f"{safe_float(entry.get('turnover_in_days_dynamic_year')):.1f}",
        ]

    def fill_row(row_data, row_num):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Окраска только динамики (столбцы 4, 5, 6)
            if col_num in (4, 5, 6):
                val = safe_float(value)
                if val > 0:
                    cell.fill = PASTEL_RED
                elif val < 0:
                    cell.fill = PASTEL_GREEN

    row_num = 2
    for item in data.get("data", []):
        row_data = create_row(item)
        fill_row(row_data, row_num)
        row_num += 1

    if sum_data := data.get("sum"):
        row_data = create_row(sum_data, label="Всего")
        fill_row(row_data, row_num)
        row_num += 1

    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def create_excel_report_for_turnover_analysis(data: Union[Dict, List[Dict]]) -> BytesIO:
    if isinstance(data, list):
        data = data[0] if data else {}

    excel_buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Оборачиваемость остатков"

    headers = [
        "Точка",
        "Остаток на \nконец \nпериода",
        "Остаток \n на конец \nпериода в днях",
        "Динамика \nнеделя",
        "Динамика \nмесяц",
        "Динамика \nгод"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    def create_row(entry: Dict, label: str = None) -> list:
        return [
            f"{label}" if label else entry.get("label", ""),
            f"{safe_float(entry.get('remainder_end')):,.0f}".replace(',', ' '),
            f"{safe_float(entry.get('turnover_in_days')):.1f}",
            f"{safe_float(entry.get('turnover_in_days_dynamic_week')):.1f}",
            f"{safe_float(entry.get('turnover_in_days_dynamic_month')):.1f}",
            f"{safe_float(entry.get('turnover_in_days_dynamic_year')):.1f}",
        ]

    def fill_row(row_data, row_num):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if col_num in (4, 5, 6):
                val = safe_float(value)
                if val > 0:
                    cell.fill = PASTEL_RED
                elif val < 0:
                    cell.fill = PASTEL_GREEN

    row_num = 2
    for item in data.get("data", []):
        row_data = create_row(item)
        fill_row(row_data, row_num)
        row_num += 1

    if sum_data := data.get("sum"):
        row_data = create_row(sum_data, label="Всего")
        fill_row(row_data, row_num)
        row_num += 1

    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
