import os
from io import BytesIO
from typing import Union, List
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Логгер
logging.basicConfig(level=logging.INFO)


def safe_float(value):
    """Преобразует значение в float или 0."""
    try:
        return float(value) if value else 0
    except ValueError:
        return 0


def format_percent(val):
    try:
        return f"{float(val):.2f}%" if val is not None else "—"
    except:
        return "—"


def write_off_parameters_create_excel_report(data: Union[List[dict], dict]) -> BytesIO:
    """Создаёт Excel отчёт по списаниям и возвращает BytesIO объект."""

    # Подготовка данных
    if isinstance(data, list):
        if not data:
            raise ValueError("Пустой список данных")
        data = data[0]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Списания по статьям"

    # Обновлённые заголовки
    headers = [
        "Статья списания", "Сумма",
        "Динамика неделя", "Динамика месяц", "Динамика год",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin"))

    row_num = 2  # Стартовая строка

    for item in data.get("data", []):
        if safe_float(item.get("write_off")) == 0:
            continue

        ws.cell(row=row_num, column=1, value=item.get("label", "—"))
        ws.cell(row=row_num, column=2, value=safe_float(item.get("write_off")))

        # Динамики — с форматированием и цветом
        for i, key in enumerate(["write_off_dynamics_week", "write_off_dynamics_month", "write_off_dynamics_year"], start=3):
            val = item.get(key)
            cell = ws.cell(row=row_num, column=i, value=format_percent(val))
            if safe_float(val) > 0:
                cell.font = Font(color="008000")  # зелёный
            elif safe_float(val) < 0:
                cell.font = Font(color="FF0000")  # красный

        row_num += 1

    # Стилизация строк
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin"))

    # Автоматическое выравнивание ширины столбцов
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
            except:
                pass
        ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    # Сохраняем в буфер
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
