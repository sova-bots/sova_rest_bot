import logging

from aiogram import Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, FSInputFile, BufferedInputFile
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Define router
foodcost_of_products_dishes_excel_router = Router()

# Helper functions (same as before)
def get_first_non_null(*values):
    for value in values:
        if value is not None:
            return value
    return "-"

def calculate_percentage_change(old_value, new_value):
    if isinstance(old_value, (int, float)) and isinstance(new_value, (int, float)) and old_value != 0:
        return round(((new_value - old_value) / old_value) * 100, 2)
    return None

def calculate_monthly_differences(store):
    diff_price = store.get("food_cost_dynamics_week")
    diff_price2 = store.get("food_cost_dynamics_month")
    diff_price3 = store.get("food_cost_dynamics_year")

    diff_price_1 = get_first_non_null(diff_price, diff_price2, diff_price3)

    if isinstance(diff_price2, (int, float)) and isinstance(diff_price, (int, float)):
        diff_price_1_week = calculate_percentage_change(diff_price, diff_price2)
    else:
        diff_price_1_week = "-"

    if isinstance(diff_price3, (int, float)) and isinstance(diff_price2, (int, float)):
        diff_price_2_month = calculate_percentage_change(diff_price2, diff_price3)
    else:
        diff_price_2_month = "-"

    return diff_price_1, diff_price_1_week, diff_price_2_month

def safe_format(value):
    if value is None:
        return "-"
    if isinstance(value, (int, float)):
        return f"{value:,.2f}%"  # Add percentage formatting
    return str(value)

# Function to create Excel report
def create_excel_report(data, output_file="food_cost_dish_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Food Cost Report"

    headers = [
        "Продукт", "Фудкость", "Динамика неделя",
        "Динамика месяц", "Динамика год"
    ]

    # Adding headers to the worksheet
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sorted_data = sorted(data["data"], key=lambda x: x["food_cost"], reverse=True)
    row_num = 2  # Start writing from the second row

    for product in sorted_data:
        food_cost_1, food_cost_1_week, food_cost_2_month = calculate_monthly_differences(product)

        row = [
            product["label"],
            safe_format(product['food_cost']),
            safe_format(food_cost_1_week),
            safe_format(food_cost_1),
            safe_format(food_cost_2_month),
        ]

        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1

    wb.save(output_file)
    print(f"Excel-файл успешно сохранен: {output_file}")

# Function to load JSON data
def load_json_data(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return data
    except Exception as e:
        print(f"Ошибка загрузки данных JSON из {file_path}: {e}")
        return None


@foodcost_of_products_dishes_excel_router.callback_query(F.data == 'format_excel_food_cost')
async def generate_excel_report_callback(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer("Формирую Excel отчёт по себестоимости блюд...")

    # Извлекаем тип отчёта из callback_data
    report_type = callback_query.data.split("_")[-1]

    # Загрузка данных из JSON-файла
    filepath = r"C:\WORK\sova_rest_bot\sova_rest_bot-master\files\jsons_for_reports\food-cost-dish_server_data_example.json"
    foodcost_data = load_json_data(filepath)  # Используем функцию для загрузки данных

    if not foodcost_data:
        await callback_query.message.answer("Ошибка при загрузке данных для отчёта.")
        return

    # Генерация Excel
    try:
        filename = f"foodcost_dishes_{report_type}.xlsx"
        create_excel_report(foodcost_data, filename)  # Используем функцию для создания Excel
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        await callback_query.message.answer(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        with open(filename, "rb") as file:
            file_data = file.read()

        input_file = BufferedInputFile(file_data, filename=f"foodcost_dishes_{report_type}.xlsx")
        await callback_query.message.answer_document(
            input_file,
            caption=f"Ваш отчёт по себестоимости блюд (тип: {report_type}):"
        )
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")
        await callback_query.message.answer("Ошибка при отправке отчёта.")
    finally:
        # Удаляем временный файл после отправки
        if os.path.exists(filename):
            os.remove(filename)
