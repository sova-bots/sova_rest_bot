from typing import Union, Dict, List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from io import BytesIO


def safe_float(value):
    try:
        return float(value) if value is not None else None
    except (ValueError, TypeError):
        return None


def loss_forecast_create_excel_report(data: Union[Dict, List[Dict]]) -> BytesIO:
    """Создаёт Excel-отчёт по прогнозу потерь по товарам."""

    if isinstance(data, list):
        if len(data) > 0 and isinstance(data[0], dict):
            data = data[0]
        else:
            raise ValueError("Ожидался список с одним словарём внутри.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Прогноз потерь"

    # Заголовки
    header = [
        "Товар",
        "Актуальная цена",
        "Последнее изменение цены",
        "Прогноз потерь / экономии"
    ]
    ws.append(header)

    # Стили
    header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    font_bold = Font(bold=True)

    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.alignment = center_align
        cell.font = font_bold

    total_actual_price = 0
    total_last_change = 0
    total_forecast = 0
    total_items = 0
    row_index = 2

    for item in data.get("data", []):
        avg_prices = [
            safe_float(item.get("avg_price_one_week_ago")),
            safe_float(item.get("avg_price_two_week_ago")),
            safe_float(item.get("avg_price_three_week_ago")),
            safe_float(item.get("avg_price_four_week_ago")),
        ]
        actual_price = next((p for p in avg_prices if p is not None), None)

        amounts = [
            safe_float(item.get("amount_one_month_ago")),
            safe_float(item.get("amount_two_month_ago")),
            safe_float(item.get("amount_three_month_ago")),
        ]
        last_change = next((a for a in amounts if a is not None), None)

        forecast = safe_float(item.get('forecast'))

        if actual_price is not None and last_change is not None:

            ws.append([
                item.get("label", ""),
                f"{actual_price:,.2f}",
                f"{last_change:,.2f}",
                f"{forecast:,.2f}"
            ])

            for col in range(1, 5):
                ws.cell(row=row_index, column=col).alignment = center_align

            total_actual_price += actual_price
            total_last_change += last_change
            total_forecast += forecast
            total_items += 1
            row_index += 1

    # Итоговая строка
    if total_items > 0:
        ws.append([
            "Итого",
            f"{total_actual_price:,.2f}",
            f"{total_last_change:,.2f}",
            f"{total_forecast:,.2f}"
        ])

        for col in range(1, 5):
            cell = ws.cell(row=row_index, column=col)
            cell.font = font_bold
            cell.fill = total_fill
            cell.alignment = center_align

    # Автоширина
    for col in range(1, len(header) + 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = 22

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def loss_forecast_only_negative_create_excel_report(data: Union[Dict, List[Dict]]) -> BytesIO:
    """Создаёт Excel-отчёт по прогнозу потерь по товарам."""

    if isinstance(data, list):
        if len(data) > 0 and isinstance(data[0], dict):
            data = data[0]
        else:
            raise ValueError("Ожидался список с одним словарём внутри.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Прогноз потерь"

    # Заголовки
    header = [
        "Товар",
        "Актуальная цена",
        "Последнее изменение цены",
        "Прогноз потерь / экономии"
    ]
    ws.append(header)

    # Стили
    header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    font_bold = Font(bold=True)

    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.alignment = center_align
        cell.font = font_bold

    total_actual_price = 0
    total_last_change = 0
    total_forecast = 0
    total_items = 0
    row_index = 2

    for item in data.get("data", []):
        avg_prices = [
            safe_float(item.get("avg_price_one_week_ago")),
            safe_float(item.get("avg_price_two_week_ago")),
            safe_float(item.get("avg_price_three_week_ago")),
            safe_float(item.get("avg_price_four_week_ago")),
        ]
        actual_price = next((p for p in avg_prices if p is not None), None)

        amounts = [
            safe_float(item.get("amount_one_month_ago")),
            safe_float(item.get("amount_two_month_ago")),
            safe_float(item.get("amount_three_month_ago")),
        ]
        last_change = next((a for a in amounts if a is not None), None)

        forecast = safe_float(item.get('forecast'))

        # Только если forecast > 0
        if actual_price is not None and last_change is not None and forecast is not None and forecast > 0:
            ws.append([
                item.get("label", ""),
                f"{actual_price:,.2f}",
                f"{last_change:,.2f}",
                f"{forecast:,.2f}"
            ])

            for col in range(1, 5):
                ws.cell(row=row_index, column=col).alignment = center_align

            total_actual_price += actual_price
            total_last_change += last_change
            total_forecast += forecast
            total_items += 1
            row_index += 1

    # Итоговая строка
    if total_items > 0:
        ws.append([
            "Итого",
            f"{total_actual_price:,.2f}",
            f"{total_last_change:,.2f}",
            f"{total_forecast:,.2f}"
        ])

        for col in range(1, 5):
            cell = ws.cell(row=row_index, column=col)
            cell.font = font_bold
            cell.fill = total_fill
            cell.alignment = center_align

    # Автоширина
    for col in range(1, len(header) + 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = 22

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

