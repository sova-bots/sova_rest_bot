import json
import logging
from aiogram import Router, types, F
from aiogram.fsm.context import FSMContext
from aiogram.types import InputFile, BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from io import BytesIO
import os

# Инициализируем роутер
trade_turnover_for_various_objects_excel_router = Router()

# Загрузка данных
def load_revenue_data(filepath: str) -> dict:
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        logging.error(f"Ошибка при чтении файла {filepath}: {e}")
        return {}

# Функция создания Excel отчета
def create_excel_report(data: dict, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Себестоимость для различных товаров"
    number_style = NamedStyle(name="number_style", number_format="#,##0")
    wb.add_named_style(number_style)

    headers = ["Заведение", "Расход в день", "Оборачиваемость (дни)", "Динамика оборачиваемости за неделю",
               "Динамика за месяц", "Динамика за год", "Оборачиваемость за неделю", "Оборачиваемость за месяц",
               "Оборачиваемость за год", "Остаток на конец периода"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in data["data"]:
        row = [
            item["label"],
            item.get("expense_day", 0),
            item.get("turnover_in_days", 0),
            item.get("turnover_in_days_dynamic_week", 0),
            item.get("turnover_in_days_dynamic_month", 0),
            item.get("turnover_in_days_dynamic_year", 0),
            item.get("turnover_in_days_week", 0),
            item.get("turnover_in_days_month", 0),
            item.get("turnover_in_days_year", 0),
            item.get("remainder_end", 0)
        ]
        ws.append(row)

    total_row = [
        data["sum"].get("label", "Всего"),
        data["sum"].get("expense_day", 0),
        data["sum"].get("turnover_in_days", 0),
        data["sum"].get("turnover_in_days_dynamic_week", 0),
        data["sum"].get("turnover_in_days_dynamic_month", 0),
        data["sum"].get("turnover_in_days_dynamic_year", 0),
        data["sum"].get("turnover_in_days_week", 0),
        data["sum"].get("turnover_in_days_month", 0),
        data["sum"].get("turnover_in_days_year", 0),
        data["sum"].get("remainder_end", 0)
    ]
    ws.append(total_row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=10):
        for cell in row:
            cell.style = number_style

    wb.save(filename)


@trade_turnover_for_various_objects_excel_router.callback_query(F.data == "format_excel_turnover_by_objects")
async def generate_excel_report_callback(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer("Формирую Excel отчёт по товарообороту...")

    # Извлекаем тип отчёта из callback_data
    report_type = callback_query.data.split("_")[-1]

    # Загрузка данных из JSON-файла
    filepath = r"C:\WORK\sova_rest_bot\sova_rest_bot-master\files\jsons_for_reports\turnouver-product_example.json"
    turnover_data = load_revenue_data(filepath)  # Используем ту же функцию для загрузки данных

    if not turnover_data:
        await callback_query.message.answer("Ошибка при загрузке данных для отчёта.")
        return

    # Генерация Excel
    try:
        filename = f"turnover_{report_type}.xlsx"
        create_excel_report(turnover_data, filename)
    except Exception as e:
        logging.error(f"Ошибка при создании Excel-файла: {e}")
        await callback_query.message.answer(f"Ошибка при создании Excel-файла: {e}")
        return

    # Отправка Excel файла
    try:
        with open(filename, "rb") as file:
            file_data = file.read()

        input_file = BufferedInputFile(file_data, filename=f"turnover_{report_type}.xlsx")
        await callback_query.message.answer_document(input_file,
                                                     caption=f"Ваш отчёт по товарообороту (тип: {report_type}):")
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel-файла: {e}")
        await callback_query.message.answer("Ошибка при отправке отчёта.")
    finally:
        # Удаляем временный файл после отправки
        if os.path.exists(filename):
            os.remove(filename)