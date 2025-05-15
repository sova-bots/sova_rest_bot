from typing import Union, Dict, List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def safe_float(value):
    try:
        return float(value) if value is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


green_fill = PatternFill(start_color="B7E4B9", end_color="B7E4B9", fill_type="solid")  # Зеленый пастельный
red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Красный пастельный
no_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Без заливки (по умолчанию)


def apply_fill_based_on_value(cell, value):
    """Применяет заливку к ячейке на основе числового значения."""
    if value < 0:
        cell.fill = green_fill
    elif value > 0:
        cell.fill = red_fill
    else:
        cell.fill = no_fill


def get_dynamic_value(item, header):
    """Извлекает значение динамики из элемента данных на основе заголовка."""
    header_key = header.lower().replace('\n', ' ').replace('динамика', '').strip().replace(' ', '_')
    if 'бар' in header_key:
        return item.get(f'food_cost_bar_{header_key.replace("_бар", "")}', 0)
    elif 'кухня' in header_key:
        return item.get(f'food_cost_kitchen_{header_key.replace("_кухня", "")}', 0)
    else:
        return item.get(f'food_cost_{header_key}', 0)


def create_excel_report_for_food_cost(data: Union[Dict, List[Dict]]) -> BytesIO:
    """Создаёт Excel-отчёт по фудкосту."""
    if isinstance(data, list):
        data = data[0] if data else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Фудкост"

    # Заголовки
    headers = [
        "Точка",
        "Фудкост", "Динамика \nнеделя", "Динамика \nмесяц", "Динамика \nгод",
        "Фудкост \nБар", "Динамика \nнеделя \nБар", "Динамика \nмесяц \nБар", "Динамика \nгод \nБар",
        "Фудкост \nКухня", "Динамика \nнеделя \nКухня", "Динамика \nмесяц \nКухня", "Динамика \nгод \nКухня"
    ]
    ws.append(headers)

    # Стили для заголовка
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Функция создания строки
    def create_row(source: Dict, label: str = None) -> List:
        return [
            label if label else source.get("label", ""),
            f"{safe_float(source.get('food_cost')):.1f}%",
            f"{safe_float(source.get('food_cost_dynamics_week')):.1f}%",
            f"{safe_float(source.get('food_cost_dynamics_month')):.1f}%",
            f"{safe_float(source.get('food_cost_dynamics_year')):.1f}%",
            f"{safe_float(source.get('food_cost_bar')):.1f}",
            f"{safe_float(source.get('food_cost_bar_dynamics_week')):.1f}%",
            f"{safe_float(source.get('food_cost_bar_dynamics_month')):.1f}%",
            f"{safe_float(source.get('food_cost_bar_dynamics_year')):.1f}%",
            f"{safe_float(source.get('food_cost_kitchen')):.1f}",
            f"{safe_float(source.get('food_cost_kitchen_dynamics_week')):.1f}%",
            f"{safe_float(source.get('food_cost_kitchen_dynamics_month')):.1f}%",
            f"{safe_float(source.get('food_cost_kitchen_dynamics_year')):.1f}%",
        ]

    # Основные строки
    for item in data.get("data", []):
        row = create_row(item)
        ws.append(row)

        # Применяем заливку к ячейкам с динамикой
        for col_num in range(2, len(headers) + 1):
            if "Динамика" in headers[col_num - 1]:
                dynamic_value = safe_float(get_dynamic_value(item, headers[col_num - 1]))
                apply_fill_based_on_value(ws.cell(row=ws.max_row, column=col_num), dynamic_value)

    # Итоговая строка "Всего"
    if sum_data := data.get("sum"):
        row = create_row(sum_data, label="Всего")
        ws.append(row)
        # Заливаем последнюю строку цветом
        last_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.font = Font(bold=True)

    # Выравнивание и ширина колонок
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    # Сохраняем в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_excel_report_for_food_cost_analysis(data: Union[Dict, List[Dict]]) -> BytesIO:
    """Создаёт Excel-отчёт по фудкост-анализу товаров и итогов."""
    if isinstance(data, list) and data:
        data = data[0]  # Берем первый элемент, если передан список

    wb = Workbook()
    ws = wb.active
    ws.title = "Фудкост"

    headers_items = [
        "Товар", "Фудкост", "Динамика неделя", "Динамика месяц", "Динамика год"
    ]

    ws.append(["1. Фудкост"])
    ws.append(headers_items)

    # Стили заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, _ in enumerate(headers_items, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные товаров
    for item in data.get("data", []):
        row = [
            item.get("label", ""),
            f"{safe_float(item.get('food_cost')):.1f}%",
            f"{safe_float(item.get('food_cost_dynamics_week')):.1f}%",
            f"{safe_float(item.get('food_cost_dynamics_month')):.1f}%",
            f"{safe_float(item.get('food_cost_dynamics_year')):.1f}%",
        ]
        ws.append(row)

        # Применяем заливку для динамики (неделя, месяц, год)
        for col_num in [3, 4, 5]:  # Колонки с динамикой
            dynamic_value = safe_float(item.get(headers_items[col_num-1].lower().replace(' ', '_')))
            apply_fill_based_on_value(ws.cell(row=ws.max_row, column=col_num), dynamic_value)

    # ========== Вторая таблица: Итог ==========
    ws.append([])  # Пустая строка
    ws.append(["2. Итоговые показатели (Всего)"])

    headers_sum = [
        " ",
        "Фудкост", "Динамика \nнеделя", "Динамика \nмесяц", "Динамика \nгод",
        "Фудкост \nБар", "Динамика \nнеделя \nБар", "Динамика \nмесяц \nБар", "Динамика \nгод \nБар",
        "Фудкост \nКухня", "Динамика \nнеделя \nКухня", "Динамика \nмесяц \nКухня", "Динамика \nгод \nКухня"
    ]
    ws.append(headers_sum)

    for col_num, _ in enumerate(headers_sum, 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sum_data = data.get("sum", {})
    row_sum = [
        str(sum_data.get("label", "Итого")),
        f"{safe_float(sum_data.get('food_cost')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_dynamics_week')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_dynamics_month')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_dynamics_year')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_bar')):.1f}",
        f"{safe_float(sum_data.get('food_cost_bar_dynamics_week')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_bar_dynamics_month')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_bar_dynamics_year')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_kitchen')):.1f}",
        f"{safe_float(sum_data.get('food_cost_kitchen_dynamics_week')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_kitchen_dynamics_month')):.1f}%",
        f"{safe_float(sum_data.get('food_cost_kitchen_dynamics_year')):.1f}%",
    ]
    ws.append(row_sum)

    # Применяем заливку к динамике в итоговой строке
    for col_num in range(2, len(headers_sum) + 1):
        if "Динамика" in headers_sum[col_num - 1]:
            dynamic_value = safe_float(get_dynamic_value(sum_data, headers_sum[col_num - 1]))
            apply_fill_based_on_value(ws.cell(row=ws.max_row, column=col_num), dynamic_value)

    # Заливаем итоговую строку цветом
    last_row = ws.max_row
    for col in range(1, len(headers_sum) + 1):
        cell = ws.cell(row=last_row, column=col)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Настройка ширины колонок
    max_cols = max(len(headers_items), len(headers_sum))
    for col_num in range(1, max_cols + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    # Сохраняем в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
