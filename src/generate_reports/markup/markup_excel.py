from typing import Union, Dict, List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def safe_float(value):
    try:
        return float(value) if value is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def create_excel_report_for_markup(data: Union[Dict, List[Dict]]) -> BytesIO:
    """Создаёт Excel-отчёт по наценке (короткая версия без периода %)."""

    if isinstance(data, list):
        data = data[0] if data else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Наценка"

    headers = [
        "Точка",
        "Наценка, %", "Динамика нед", "Динамика мес", "Динамика год"
    ]
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")

    # Стили для значений
    positive_fill = PatternFill(start_color="D0F0C0", end_color="D0F0C0", fill_type="solid")  # Светло-зеленый
    negative_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Светло-красный

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def create_row(source: Dict, label: str = None) -> List:
        return [
            label if label else source.get("label", ""),
            safe_float(source.get('markup')),
            safe_float(source.get('markup_dynamics_week')),
            safe_float(source.get('markup_dynamics_month')),
            safe_float(source.get('markup_dynamics_year')),
        ]

    for item in data.get("data", []):
        row_values = create_row(item)
        ws.append([
            row_values[0],  # Название точки
            f"{row_values[1]:.1f}%",
            f"{row_values[2]:+.1f}%",
            f"{row_values[3]:+.1f}%",
            f"{row_values[4]:+.1f}%",
        ])

        # Применяем цветовую схему к числовым ячейкам
        current_row = ws.max_row
        for col_num in [2, 3, 4, 5]:  # Колонки с числами
            value = row_values[col_num - 1]
            cell = ws.cell(row=current_row, column=col_num)

            if value > 0:
                cell.fill = positive_fill
            elif value < 0:
                cell.fill = negative_fill

    # Настройка ширины столбцов и выравнивания
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_excel_report_for_markup_analysis(data: Union[Dict, List[Dict]]) -> BytesIO:
    """Создаёт Excel-отчёт по анализу наценки (короткая версия без периода %)."""

    if isinstance(data, list):
        data = data[0] if data else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ наценки"

    headers = [
        "Товар",
        "Наценка, %", "Динамика нед", "Динамика мес", "Динамика год"
    ]
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")

    # Стили для значений
    positive_fill = PatternFill(start_color="D0F0C0", end_color="D0F0C0", fill_type="solid")  # Светло-зеленый
    negative_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Светло-красный

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def create_row(source: Dict, label: str = None) -> List:
        return [
            label if label else source.get("label", ""),
            safe_float(source.get('markup')),
            safe_float(source.get('markup_dynamics_week')),
            safe_float(source.get('markup_dynamics_month')),
            safe_float(source.get('markup_dynamics_year')),
        ]

    for item in data.get("data", []):
        row_values = create_row(item)
        ws.append([
            row_values[0],  # Название товара
            f"{row_values[1]:.1f}%",
            f"{row_values[2]:+.1f}%",
            f"{row_values[3]:+.1f}%",
            f"{row_values[4]:+.1f}%",
        ])

        # Применяем цветовую схему к числовым ячейкам
        current_row = ws.max_row
        for col_num in [2, 3, 4, 5]:  # Колонки с числами
            value = row_values[col_num - 1]
            cell = ws.cell(row=current_row, column=col_num)

            if value > 0:
                cell.fill = positive_fill
            elif value < 0:
                cell.fill = negative_fill


    # Настройка ширины столбцов и выравнивания
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
